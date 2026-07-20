#!/usr/bin/env python3
# DENBA Max 進銷存 — single-file webapp (Flask + SQLite + waitress), multi-user
import datetime
import io
import json
import os
import re
import secrets
import sqlite3
import sys
import time
from functools import wraps

from flask import Flask, g, jsonify, request, send_file, session
from werkzeug.security import check_password_hash, generate_password_hash
from webauthn import (
    base64url_to_bytes,
    generate_authentication_options,
    generate_registration_options,
    options_to_json,
    verify_authentication_response,
    verify_registration_response,
)
from webauthn.helpers import bytes_to_base64url
from webauthn.helpers.exceptions import InvalidAuthenticationResponse, InvalidRegistrationResponse
from webauthn.helpers.structs import (
    AttestationConveyancePreference,
    AuthenticatorSelectionCriteria,
    PublicKeyCredentialDescriptor,
    ResidentKeyRequirement,
    UserVerificationRequirement,
)

BASE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("DB_PATH", os.path.join(BASE, "denba.db"))
PORT = int(os.environ.get("PORT", "2026"))
APP_PASSWORD = os.environ.get("APP_PASSWORD", "")   # bootstrap admin password (first run only)
APP_USER = os.environ.get("APP_USER", "admin")      # bootstrap admin username (first run only)
BACKUP_DIR = os.environ.get("BACKUP_DIR", os.path.join(os.path.dirname(DB_PATH), "backups"))
RP_ID = os.environ.get("RP_ID", "denba.hydr0negnetwork.de")
ORIGIN = os.environ.get("ORIGIN", "https://denba.hydr0negnetwork.de")
RP_NAME = "DENBA 進銷存"

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
app.config.update(
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_HTTPONLY=True,
    MAX_CONTENT_LENGTH=1024 * 1024,
    PERMANENT_SESSION_LIFETIME=datetime.timedelta(minutes=30),
)


@app.after_request
def security_headers(resp):
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "DENY")
    resp.headers.setdefault("Referrer-Policy", "same-origin")
    if request.path.startswith("/api/"):
        resp.headers["Cache-Control"] = "no-store"
    return resp


@app.before_request
def limit_field_lengths():
    if request.method in ("POST", "PATCH") and request.path.startswith("/api/"):
        d = request.get_json(silent=True)
        def walk(v):
            if isinstance(v, str):
                if len(v) > 1000:
                    return False
            elif isinstance(v, dict):
                for val in v.values():
                    if not walk(val):
                        return False
            elif isinstance(v, list):
                for val in v:
                    if not walk(val):
                        return False
            return True
        if d is not None and not isinstance(d, dict):
            return bad("資料格式不正確")
        if d is not None and not walk(d):
            return bad("欄位長度過長")

SCHEMA = """
CREATE TABLE IF NOT EXISTS users(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  username TEXT NOT NULL UNIQUE,
  password_hash TEXT NOT NULL,
  is_admin INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE IF NOT EXISTS purchases(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  date TEXT NOT NULL,
  model TEXT NOT NULL,
  qty INTEGER NOT NULL,
  total INTEGER NOT NULL,
  note TEXT NOT NULL DEFAULT '',
  user_id INTEGER NOT NULL DEFAULT 1
);
CREATE TABLE IF NOT EXISTS units(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  serial TEXT NOT NULL,
  model TEXT NOT NULL,
  purchase_id INTEGER,
  cost INTEGER NOT NULL,
  status TEXT NOT NULL DEFAULT 'in_stock',
  note TEXT NOT NULL DEFAULT '',
  user_id INTEGER NOT NULL DEFAULT 1
);
CREATE TABLE IF NOT EXISTS sales(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  date TEXT NOT NULL,
  customer TEXT NOT NULL,
  unit_id INTEGER,
  model TEXT NOT NULL,
  serial TEXT NOT NULL DEFAULT '',
  price INTEGER NOT NULL,
  card_fee INTEGER NOT NULL DEFAULT 0,
  cost INTEGER NOT NULL,
  warranty_no TEXT NOT NULL DEFAULT '',
  note TEXT NOT NULL DEFAULT '',
  user_id INTEGER NOT NULL DEFAULT 1,
  sale_type TEXT NOT NULL DEFAULT 'normal',
  agent TEXT NOT NULL DEFAULT '',
  deposit INTEGER NOT NULL DEFAULT 0,
  deposit_date TEXT NOT NULL DEFAULT '',
  commission INTEGER NOT NULL DEFAULT 0,
  tax INTEGER NOT NULL DEFAULT 0,
  health_fee INTEGER NOT NULL DEFAULT 0,
  settled INTEGER NOT NULL DEFAULT 0,
  settle_date TEXT NOT NULL DEFAULT '',
  extra_fee INTEGER NOT NULL DEFAULT 0,
  extra_label TEXT NOT NULL DEFAULT '',
  group_id INTEGER
);
CREATE TABLE IF NOT EXISTS trials(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  customer TEXT NOT NULL,
  model TEXT NOT NULL,
  start_date TEXT NOT NULL DEFAULT '',
  end_date TEXT NOT NULL DEFAULT '',
  note TEXT NOT NULL DEFAULT '',
  returned INTEGER NOT NULL DEFAULT 0,
  user_id INTEGER NOT NULL DEFAULT 1,
  rent_type TEXT NOT NULL DEFAULT '',
  return_date TEXT NOT NULL DEFAULT ''
);
CREATE TABLE IF NOT EXISTS consignments(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  agent TEXT NOT NULL,
  unit_id INTEGER,
  deposit INTEGER NOT NULL DEFAULT 0,
  deposit_date TEXT NOT NULL DEFAULT '',
  note TEXT NOT NULL DEFAULT '',
  user_id INTEGER NOT NULL DEFAULT 1,
  returned INTEGER NOT NULL DEFAULT 0,
  refund_date TEXT NOT NULL DEFAULT '',
  refund_amount INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE IF NOT EXISTS webauthn_credentials(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  user_id INTEGER NOT NULL,
  credential_id TEXT NOT NULL UNIQUE,
  public_key TEXT NOT NULL,
  sign_count INTEGER NOT NULL DEFAULT 0,
  transports TEXT NOT NULL DEFAULT '',
  label TEXT NOT NULL DEFAULT '',
  created TEXT NOT NULL DEFAULT ''
);
"""

UNIT_STATUSES = ("in_stock", "sold", "trial", "retired", "consigned")
RENT_TYPES = ("week7", "month", "franchise", "hq", "reserve")
DATA_TABLES = ("purchases", "units", "sales", "trials", "consignments")
USERNAME_RE = re.compile(r"^[A-Za-z0-9_.-]{2,20}$")
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
WITHHOLD_RATE = 0.10   # 預扣稅款
HEALTH_RATE = 0.0211   # 二代健保補充保費


def valid_date(s):
    if not DATE_RE.match(s):
        return False
    try:
        datetime.date.fromisoformat(s)
    except ValueError:
        return False
    return True


def half_up(x):
    return int(x + 0.5)   # Taiwan 四捨五入 — Python round() is banker's rounding, don't use it here


def next_month_15(date_s):
    d = datetime.date.fromisoformat(date_s)
    y, m = (d.year + 1, 1) if d.month == 12 else (d.year, d.month + 1)
    return f"{y:04d}-{m:02d}-15"


def init_db():
    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA busy_timeout=5000")
    con.execute("PRAGMA journal_mode=WAL")
    con.executescript(SCHEMA)
    cols = [r[1] for r in con.execute("PRAGMA table_info(sales)")]
    if "warranty_no" not in cols:
        con.execute("ALTER TABLE sales ADD COLUMN warranty_no TEXT NOT NULL DEFAULT ''")
    if "group_id" not in cols:
        con.execute("ALTER TABLE sales ADD COLUMN group_id INTEGER")
    franchise_cols = {
        "sale_type": "TEXT NOT NULL DEFAULT 'normal'",
        "agent": "TEXT NOT NULL DEFAULT ''",
        "deposit": "INTEGER NOT NULL DEFAULT 0",
        "deposit_date": "TEXT NOT NULL DEFAULT ''",
        "commission": "INTEGER NOT NULL DEFAULT 0",
        "tax": "INTEGER NOT NULL DEFAULT 0",
        "health_fee": "INTEGER NOT NULL DEFAULT 0",
        "settled": "INTEGER NOT NULL DEFAULT 0",
        "settle_date": "TEXT NOT NULL DEFAULT ''",
        "extra_fee": "INTEGER NOT NULL DEFAULT 0",
        "extra_label": "TEXT NOT NULL DEFAULT ''",
    }
    for col, ddl in franchise_cols.items():
        if col not in cols:
            con.execute(f"ALTER TABLE sales ADD COLUMN {col} {ddl}")
    ccols = [r[1] for r in con.execute("PRAGMA table_info(consignments)")]
    consign_cols = {
        "returned": "INTEGER NOT NULL DEFAULT 0",
        "refund_date": "TEXT NOT NULL DEFAULT ''",
        "refund_amount": "INTEGER NOT NULL DEFAULT 0",
    }
    for col, ddl in consign_cols.items():
        if col not in ccols:
            con.execute(f"ALTER TABLE consignments ADD COLUMN {col} {ddl}")
    tcols = [r[1] for r in con.execute("PRAGMA table_info(trials)")]
    if "rent_type" not in tcols:
        con.execute("ALTER TABLE trials ADD COLUMN rent_type TEXT NOT NULL DEFAULT ''")
    if "return_date" not in tcols:
        con.execute("ALTER TABLE trials ADD COLUMN return_date TEXT NOT NULL DEFAULT ''")
    ucols = [r[1] for r in con.execute("PRAGMA table_info(users)")]
    if "token_ver" not in ucols:
        con.execute("ALTER TABLE users ADD COLUMN token_ver INTEGER NOT NULL DEFAULT 0")
    if "shares_with" not in ucols:
        con.execute("ALTER TABLE users ADD COLUMN shares_with INTEGER")
    # multi-user migration: add user_id to legacy tables (existing rows → user 1)
    for t in DATA_TABLES:
        tcols = [r[1] for r in con.execute(f"PRAGMA table_info({t})")]
        if "user_id" not in tcols:
            con.execute(f"ALTER TABLE {t} ADD COLUMN user_id INTEGER NOT NULL DEFAULT 1")
    # legacy units had a global UNIQUE(serial); rebuild so serials are unique per user
    units_sql = con.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='units'").fetchone()[0]
    if "UNIQUE" in units_sql.upper():
        con.executescript("""
            CREATE TABLE units_mu(
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              serial TEXT NOT NULL,
              model TEXT NOT NULL,
              purchase_id INTEGER,
              cost INTEGER NOT NULL,
              status TEXT NOT NULL DEFAULT 'in_stock',
              note TEXT NOT NULL DEFAULT '',
              user_id INTEGER NOT NULL DEFAULT 1
            );
            INSERT INTO units_mu(id,serial,model,purchase_id,cost,status,note,user_id)
              SELECT id,serial,model,purchase_id,cost,status,note,user_id FROM units;
            DROP TABLE units;
            ALTER TABLE units_mu RENAME TO units;
        """)
    con.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_units_user_serial ON units(user_id, serial)")
    # v29 lookup indexes (every query is user-scoped; units are also joined by purchase, sales by unit)
    con.execute("CREATE INDEX IF NOT EXISTS idx_units_purchase ON units(purchase_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_sales_user ON sales(user_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_sales_unit ON sales(unit_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_sales_group ON sales(user_id, group_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_purchases_user ON purchases(user_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_trials_user ON trials(user_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_consign_user ON consignments(user_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_webauthn_user ON webauthn_credentials(user_id)")
    # bootstrap admin on first run
    if con.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0 and APP_PASSWORD:
        con.execute(
            "INSERT INTO users(username,password_hash,is_admin) VALUES(?,?,1)",
            (APP_USER, generate_password_hash(APP_PASSWORD)),
        )
    empty = all(
        con.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0] == 0 for t in DATA_TABLES
    )
    seed_path = os.path.join(BASE, "seed_data.json")
    if empty and os.path.exists(seed_path):
        with open(seed_path, encoding="utf-8") as f:
            seed = json.load(f)
        for p in seed.get("purchases", []):
            con.execute(
                "INSERT INTO purchases(date,model,qty,total,note,user_id) VALUES(?,?,?,?,?,1)",
                (p["date"], p["model"], p["qty"], p["total"], p.get("note", "")),
            )
        for u in seed.get("units", []):
            con.execute(
                "INSERT INTO units(serial,model,purchase_id,cost,status,note,user_id)"
                " VALUES(?,?,NULL,?,?,?,1)",
                (u["serial"], u["model"], u["cost"], u["status"], u.get("note", "")),
            )
        for s in seed.get("sales", []):
            con.execute(
                "INSERT INTO sales(date,customer,unit_id,model,serial,price,card_fee,cost,note,user_id)"
                " VALUES(?,?,NULL,?,?,?,?,?,?,1)",
                (s["date"], s["customer"], s["model"], s.get("serial", ""),
                 s["price"], s.get("card_fee", 0), s["cost"], s.get("note", "")),
            )
        for t in seed.get("trials", []):
            con.execute(
                "INSERT INTO trials(customer,model,start_date,end_date,note,returned,user_id,rent_type,return_date)"
                " VALUES(?,?,?,?,?,?,1,?,?)",
                (t.get("customer", ""), t.get("model", ""), t.get("start_date", ""),
                 t.get("end_date", ""), t.get("note", ""), t.get("returned", 0),
                 t.get("rent_type", ""), t.get("return_date", "")),
            )
    con.commit()
    con.close()


def db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.execute("PRAGMA busy_timeout=5000")
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(_exc):
    con = g.pop("db", None)
    if con is not None:
        con.close()


def auth_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        uid = session.get("uid")
        if not uid:
            return jsonify(error="unauthorized"), 401
        user = db().execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        if not user or session.get("tv", 0) != user["token_ver"]:
            session.clear()
            return jsonify(error="unauthorized"), 401
        g.user = user
        g.data_uid = user["shares_with"] or user["id"]
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    @auth_required
    def wrapper(*args, **kwargs):
        if not g.user["is_admin"]:
            return jsonify(error="需要管理員權限"), 403
        return f(*args, **kwargs)
    return wrapper


def bad(msg, code=400):
    return jsonify(error=msg), code


def as_int(v, default=0):
    try:
        return int(round(float(v)))
    except (TypeError, ValueError, OverflowError):
        return default


# ---------- pages ----------

@app.route("/")
def index():
    return app.send_static_file("index.html")


@app.route("/sw.js")
def sw():
    return app.send_static_file("sw.js")


# ---------- auth ----------

# login rate limiting: per-IP and per-username failure lockout (in-memory)
LOGIN_FAILS = {}          # key -> (count, first_fail_ts, locked_until_ts)
IP_LIMIT, USER_LIMIT = 10, 20
LOCK_SECS = 900
FAIL_WINDOW = 900
DUMMY_HASH = generate_password_hash("timing-equalizer")


def client_ip():
    return (request.headers.get("CF-Connecting-IP")
            or request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
            or request.remote_addr or "?")


def login_locked(keys):
    now = time.time()
    for k in list(LOGIN_FAILS):
        c, first, locked = LOGIN_FAILS[k]
        if locked < now and now - first > FAIL_WINDOW:
            LOGIN_FAILS.pop(k, None)
    for k in keys:
        rec = LOGIN_FAILS.get(k)
        if rec and rec[2] > now:
            return int(rec[2] - now)
    return 0


def login_failed(keys):
    now = time.time()
    for k in keys:
        limit = IP_LIMIT if k.startswith("ip:") else USER_LIMIT
        c, first, locked = LOGIN_FAILS.get(k, (0, now, 0.0))
        if now - first > FAIL_WINDOW:
            c, first = 0, now
        c += 1
        if c >= limit:
            locked = now + LOCK_SECS
        LOGIN_FAILS[k] = (c, first, locked)


@app.route("/api/login", methods=["POST"])
def login():
    d = request.get_json(silent=True) or {}
    username = (d.get("username") or "").strip()
    pw = d.get("password", "")
    keys = ["ip:" + client_ip(), "u:" + username.lower()]
    wait = login_locked(keys)
    if wait:
        return bad(f"嘗試次數過多，請 {wait // 60 + 1} 分鐘後再試", 429)
    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA busy_timeout=5000")
    con.row_factory = sqlite3.Row
    user = con.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    con.close()
    if user:
        pw_ok = check_password_hash(user["password_hash"], pw)
    else:
        check_password_hash(DUMMY_HASH, pw)   # equalize timing; no username oracle
        pw_ok = False
    if pw_ok:
        for k in keys:
            LOGIN_FAILS.pop(k, None)
        session.permanent = True
        session["uid"] = user["id"]
        session["tv"] = user["token_ver"]
        return jsonify(ok=True, username=user["username"], is_admin=bool(user["is_admin"]))
    login_failed(keys)
    time.sleep(0.6)
    return bad("帳號或密碼錯誤", 401)


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify(ok=True)


@app.route("/api/me/password", methods=["POST"])
@auth_required
def change_own_password():
    d = request.get_json(silent=True) or {}
    old, new = d.get("old", ""), d.get("new", "")
    if len(new) < 8:
        return bad("新密碼至少 8 碼")
    if not check_password_hash(g.user["password_hash"], old):
        time.sleep(0.6)
        return bad("目前密碼錯誤")
    con = db()
    new_ver = g.user["token_ver"] + 1
    con.execute("UPDATE users SET password_hash=?, token_ver=? WHERE id=?",
                (generate_password_hash(new), new_ver, g.user["id"]))
    if d.get("reset_bio"):
        con.execute("DELETE FROM webauthn_credentials WHERE user_id=?", (g.user["id"],))
    con.commit()
    session["tv"] = new_ver
    return jsonify(ok=True)


# ---------- webauthn ----------

@app.route("/api/webauthn/register/begin", methods=["POST"])
@auth_required
def webauthn_register_begin():
    con = db()
    existing = con.execute(
        "SELECT credential_id FROM webauthn_credentials WHERE user_id=?", (g.user["id"],)
    ).fetchall()
    options = generate_registration_options(
        rp_id=RP_ID,
        rp_name=RP_NAME,
        user_id=str(g.user["id"]).encode(),
        user_name=g.user["username"],
        user_display_name=g.user["username"],
        attestation=AttestationConveyancePreference.NONE,
        authenticator_selection=AuthenticatorSelectionCriteria(
            resident_key=ResidentKeyRequirement.REQUIRED,
            user_verification=UserVerificationRequirement.REQUIRED,
        ),
        exclude_credentials=[
            PublicKeyCredentialDescriptor(id=base64url_to_bytes(r["credential_id"]))
            for r in existing
        ],
    )
    session["webauthn_challenge"] = bytes_to_base64url(options.challenge)
    return options_to_json(options), 200, {"Content-Type": "application/json"}


@app.route("/api/webauthn/register/complete", methods=["POST"])
@auth_required
def webauthn_register_complete():
    challenge = session.pop("webauthn_challenge", None)
    if not challenge:
        return bad("註冊逾時，請重試")
    body = request.get_json(silent=True) or {}
    credential = body.get("credential") or body
    label = (body.get("label") or "").strip()[:100]
    try:
        verification = verify_registration_response(
            credential=credential,
            expected_challenge=base64url_to_bytes(challenge),
            expected_rp_id=RP_ID,
            expected_origin=ORIGIN,
            require_user_verification=True,
        )
    except InvalidRegistrationResponse:
        return bad("驗證失敗")
    transports = (credential.get("response") or {}).get("transports") or []
    con = db()
    try:
        con.execute(
            "INSERT INTO webauthn_credentials(user_id,credential_id,public_key,sign_count,transports,label,created)"
            " VALUES(?,?,?,?,?,?,?)",
            (g.user["id"], bytes_to_base64url(verification.credential_id),
             bytes_to_base64url(verification.credential_public_key), verification.sign_count,
             json.dumps(transports), label, datetime.date.today().isoformat()),
        )
    except sqlite3.IntegrityError:
        return bad("此金鑰已註冊過")
    con.commit()
    return jsonify(ok=True)


@app.route("/api/webauthn/status")
def webauthn_status():
    n = db().execute("SELECT COUNT(*) FROM webauthn_credentials").fetchone()[0]
    return jsonify(available=n > 0)


@app.route("/api/webauthn/login/begin", methods=["POST"])
def webauthn_login_begin():
    ip_keys = ["ip:" + client_ip()]
    wait = login_locked(ip_keys)
    if wait:
        return bad(f"嘗試次數過多，請 {wait // 60 + 1} 分鐘後再試", 429)
    options = generate_authentication_options(
        rp_id=RP_ID,
        user_verification=UserVerificationRequirement.REQUIRED,
    )
    session["webauthn_challenge"] = bytes_to_base64url(options.challenge)
    return options_to_json(options), 200, {"Content-Type": "application/json"}


@app.route("/api/webauthn/login/complete", methods=["POST"])
def webauthn_login_complete():
    ip_keys = ["ip:" + client_ip()]
    wait = login_locked(ip_keys)
    if wait:
        return bad(f"嘗試次數過多，請 {wait // 60 + 1} 分鐘後再試", 429)
    challenge = session.pop("webauthn_challenge", None)
    if not challenge:
        login_failed(ip_keys)
        return bad("登入逾時，請重試")
    body = request.get_json(silent=True) or {}
    credential = body.get("credential") or body
    handle_b64 = (credential.get("response") or {}).get("userHandle")
    if not handle_b64:
        login_failed(ip_keys)
        return bad("無效的憑證")
    try:
        uid = int(base64url_to_bytes(handle_b64).decode("utf-8"))
    except (UnicodeDecodeError, ValueError):
        login_failed(ip_keys)
        return bad("無效的憑證")
    con = db()
    user = con.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not user:
        login_failed(ip_keys)
        return bad("無效的憑證")
    cred_id = credential.get("id")
    row = con.execute(
        "SELECT * FROM webauthn_credentials WHERE credential_id=? AND user_id=?",
        (cred_id, uid),
    ).fetchone()
    if not row:
        login_failed(ip_keys)
        return bad("無效的憑證")
    try:
        verification = verify_authentication_response(
            credential=credential,
            expected_challenge=base64url_to_bytes(challenge),
            expected_rp_id=RP_ID,
            expected_origin=ORIGIN,
            credential_public_key=base64url_to_bytes(row["public_key"]),
            credential_current_sign_count=row["sign_count"],
            require_user_verification=True,
        )
    except InvalidAuthenticationResponse:
        login_failed(ip_keys)
        return bad("驗證失敗")
    if verification.new_sign_count > row["sign_count"]:
        con.execute("UPDATE webauthn_credentials SET sign_count=? WHERE id=?",
                    (verification.new_sign_count, row["id"]))
    for k in ip_keys:
        LOGIN_FAILS.pop(k, None)
    con.commit()
    session.permanent = True
    session["uid"] = user["id"]
    session["tv"] = user["token_ver"]
    return jsonify(ok=True, username=user["username"], is_admin=bool(user["is_admin"]))


@app.route("/api/webauthn/credentials")
@auth_required
def webauthn_list_credentials():
    rows = db().execute(
        "SELECT id, label, created FROM webauthn_credentials WHERE user_id=? ORDER BY id",
        (g.user["id"],),
    ).fetchall()
    return jsonify(credentials=[
        {"id": r["id"], "label": r["label"], "created": r["created"]} for r in rows
    ])


@app.route("/api/webauthn/credentials/<int:cid>", methods=["DELETE"])
@auth_required
def webauthn_delete_credential(cid):
    con = db()
    con.execute("DELETE FROM webauthn_credentials WHERE id=? AND user_id=?", (cid, g.user["id"]))
    con.commit()
    return jsonify(ok=True)


# ---------- user management (admin) ----------

@app.route("/api/users")
@admin_required
def list_users():
    con = db()
    out = []
    for u in con.execute("SELECT id, username, is_admin, shares_with FROM users ORDER BY id"):
        counts = {t: con.execute(
            f"SELECT COUNT(*) FROM {t} WHERE user_id=?", (u["id"],)).fetchone()[0]
            for t in DATA_TABLES}
        out.append({"id": u["id"], "username": u["username"],
                    "is_admin": bool(u["is_admin"]), "shares_with": u["shares_with"],
                    "counts": counts})
    return jsonify(users=out)


@app.route("/api/users", methods=["POST"])
@admin_required
def create_user():
    d = request.get_json(silent=True) or {}
    username = (d.get("username") or "").strip()
    password = d.get("password", "")
    is_admin = 1 if d.get("is_admin") else 0
    if not USERNAME_RE.match(username):
        return bad("帳號限 2–20 位英數字（可含 _ . -）")
    if len(password) < 8:
        return bad("密碼至少 8 碼")
    con = db()
    if con.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
        return bad("帳號已存在")
    con.execute("INSERT INTO users(username,password_hash,is_admin) VALUES(?,?,?)",
                (username, generate_password_hash(password), is_admin))
    con.commit()
    return jsonify(ok=True)


@app.route("/api/users/<int:target>", methods=["PATCH"])
@admin_required
def edit_user(target):
    d = request.get_json(silent=True) or {}
    con = db()
    u = con.execute("SELECT * FROM users WHERE id=?", (target,)).fetchone()
    if not u:
        return bad("找不到使用者", 404)
    if "password" in d:
        if len(d["password"]) < 8:
            return bad("密碼至少 8 碼")
        con.execute("UPDATE users SET password_hash=?, token_ver=token_ver+1 WHERE id=?",
                    (generate_password_hash(d["password"]), target))
    if "is_admin" in d:
        new_admin = 1 if d["is_admin"] else 0
        if u["is_admin"] and not new_admin:
            admins = con.execute("SELECT COUNT(*) FROM users WHERE is_admin=1").fetchone()[0]
            if admins <= 1:
                return bad("至少需保留一位管理員")
        con.execute("UPDATE users SET is_admin=? WHERE id=?", (new_admin, target))
    if "shares_with" in d:
        new_shares_with = as_int(d["shares_with"], 0) or None
        if new_shares_with:
            if new_shares_with == target:
                return bad("無法與自己共用資料")
            share_target = con.execute(
                "SELECT shares_with FROM users WHERE id=?", (new_shares_with,)).fetchone()
            if not share_target:
                return bad("找不到共用對象", 404)
            if share_target["shares_with"]:
                return bad("共用對象本身已在共用他人資料", 400)
            con.execute("UPDATE users SET shares_with=? WHERE id=?", (new_shares_with, target))
        else:
            con.execute("UPDATE users SET shares_with=NULL WHERE id=?", (target,))
    con.commit()
    return jsonify(ok=True)


@app.route("/api/users/<int:target>", methods=["DELETE"])
@admin_required
def delete_user(target):
    con = db()
    u = con.execute("SELECT * FROM users WHERE id=?", (target,)).fetchone()
    if not u:
        return bad("找不到使用者", 404)
    if target == g.user["id"]:
        return bad("無法刪除自己")
    sharer = con.execute("SELECT id FROM users WHERE shares_with=?", (target,)).fetchone()
    if sharer:
        return bad("無法刪除：其他使用者正共用此帳號的資料", 400)
    keep = write_user_snapshot(con, target, "pre-delete")
    for fn in os.listdir(BACKUP_DIR):
        m = USER_BK_RE.match(fn)
        if m and m.group(1) == str(target) and fn != keep:
            os.unlink(os.path.join(BACKUP_DIR, fn))
    for t in DATA_TABLES:
        con.execute(f"DELETE FROM {t} WHERE user_id=?", (target,))
    con.execute("DELETE FROM users WHERE id=?", (target,))
    con.commit()
    return jsonify(ok=True)


# ---------- data ----------

MONTHLY_SQL = """
SELECT substr(date,1,7) AS ym,
       SUM(price - card_fee) AS revenue,
       SUM(cost) AS cost,
       SUM(commission) AS commission,
       SUM(extra_fee) AS extra,
       SUM(price - card_fee) - SUM(cost) - SUM(commission) - SUM(extra_fee) AS profit,
       COUNT(*) AS qty
FROM sales WHERE user_id=? GROUP BY ym ORDER BY ym
"""

# cash-basis view of 居間特許 money flow: deposit receipt counts in the deposit month
# (whether the machine has sold yet — consignments — or not — the sale row),
# deposit refund + net commission payout count in the settlement month
FRANCHISE_FLOW_SQL = """
SELECT ym, SUM(dep_in) AS dep_in, SUM(dep_out) AS dep_out,
       SUM(comm_net) AS comm_net, SUM(tax) AS tax, SUM(health) AS health
FROM (
  SELECT substr(deposit_date,1,7) AS ym, deposit AS dep_in, 0 AS dep_out, 0 AS comm_net, 0 AS tax, 0 AS health
    FROM sales WHERE user_id=? AND sale_type='franchise' AND deposit_date<>'' AND deposit>0
  UNION ALL
  SELECT substr(settle_date,1,7), 0, deposit, commission - tax - health_fee, tax, health_fee
    FROM sales WHERE user_id=? AND sale_type='franchise' AND settled=1 AND settle_date<>''
  UNION ALL
  SELECT substr(deposit_date,1,7), deposit, 0, 0, 0, 0
    FROM consignments WHERE user_id=? AND deposit_date<>'' AND deposit>0
  UNION ALL
  SELECT substr(refund_date,1,7), 0, refund_amount, 0, 0, 0
    FROM consignments WHERE user_id=? AND returned=1 AND refund_date<>'' AND refund_amount>0
) GROUP BY ym ORDER BY ym
"""


@app.route("/api/data")
@auth_required
def data():
    con = db()
    uid = g.data_uid
    return jsonify(
        me={"username": g.user["username"], "is_admin": bool(g.user["is_admin"])},
        units=[dict(r) for r in con.execute(
            "SELECT * FROM units WHERE user_id=? ORDER BY model, serial", (uid,))],
        purchases=[dict(r) for r in con.execute(
            "SELECT * FROM purchases WHERE user_id=? ORDER BY date DESC, id DESC", (uid,))],
        sales=[dict(r) for r in con.execute(
            "SELECT * FROM sales WHERE user_id=? ORDER BY date DESC, id DESC", (uid,))],
        trials=[dict(r) for r in con.execute(
            "SELECT * FROM trials WHERE user_id=? ORDER BY returned, start_date DESC, id DESC", (uid,))],
        monthly=[dict(r) for r in con.execute(MONTHLY_SQL, (uid,))],
        franchise_flow=[dict(r) for r in con.execute(FRANCHISE_FLOW_SQL, (uid, uid, uid, uid))],
        consignments=[dict(r) for r in con.execute(
            "SELECT c.*, u.serial AS serial, u.model AS model FROM consignments c"
            " LEFT JOIN units u ON u.id=c.unit_id WHERE c.user_id=?"
            " ORDER BY c.deposit_date DESC, c.id DESC", (uid,))],
        agent_owing=[dict(r) for r in con.execute(
            "SELECT agent, COUNT(*) AS n,"
            " SUM(deposit) AS deposit, SUM(commission - tax - health_fee) AS net_comm,"
            " SUM(deposit + commission - tax - health_fee) AS payable"
            " FROM sales WHERE user_id=? AND sale_type='franchise' AND settled=0 AND (deposit>0 OR commission>0) AND agent<>''"
            " GROUP BY agent ORDER BY agent", (uid,))],
        customers=[r[0] for r in con.execute(
            "SELECT DISTINCT customer FROM sales WHERE user_id=? AND customer<>'' ORDER BY customer", (uid,))],
        agents=[r[0] for r in con.execute(
            "SELECT DISTINCT agent FROM sales WHERE user_id=? AND agent<>'' ORDER BY agent", (uid,))],
    )


# ---------- purchases ----------

@app.route("/api/purchase", methods=["POST"])
@auth_required
def add_purchase():
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    date = (d.get("date") or "").strip()
    status = d.get("status", "in_stock")
    note = d.get("note", "")

    if not date:
        return bad("日期、型號、金額、貨號皆為必填")
    if not valid_date(date):
        return bad("日期格式須為 YYYY-MM-DD")
    if status not in ("in_stock", "trial"):
        return bad("入庫類型不正確")

    if "items" in d:
        items = d.get("items")
    else:
        items = [{
            "model": d.get("model"),
            "total": d.get("total"),
            "serials": d.get("serials")
        }]

    if not isinstance(items, list) or not (1 <= len(items) <= 10):
        return bad("型號筆數不正確")

    validated_items = []
    all_serials = []
    for item in items:
        if not isinstance(item, dict):
            return bad("日期、型號、金額、貨號皆為必填")
        item_model = (item.get("model") or "").strip()
        item_total = as_int(item.get("total"), -1)
        raw_serials = item.get("serials")
        if not item_model or item_total < 0 or not isinstance(raw_serials, list):
            return bad("日期、型號、金額、貨號皆為必填")
        
        item_serials = []
        for s in raw_serials:
            if not isinstance(s, str) or not s.strip():
                return bad("日期、型號、金額、貨號皆為必填")
            item_serials.append(s.strip())
            
        n = len(item_serials)
        if n == 0:
            return bad("日期、型號、金額、貨號皆為必填")
        if n > 50:
            return bad("一次最多 50 台")
            
        validated_items.append({
            "model": item_model,
            "total": item_total,
            "serials": item_serials
        })
        all_serials.extend(item_serials)

    # Check duplicates within the payload
    seen = set()
    dups = []
    for s in all_serials:
        if s in seen:
            if s not in dups:
                dups.append(s)
        else:
            seen.add(s)
    if dups:
        return bad("貨號重複：" + "、".join(dups))

    con = db()
    con.execute("BEGIN IMMEDIATE")
    try:
        # Check against existing database units
        exists = [s for s in all_serials if con.execute(
            "SELECT 1 FROM units WHERE serial=? AND user_id=?", (s, uid)).fetchone()]
        if exists:
            con.rollback()
            return bad("貨號已存在：" + "、".join(exists))

        pids = []
        for item in validated_items:
            item_model = item["model"]
            item_total = item["total"]
            item_serials = item["serials"]
            n = len(item_serials)
            base_cost = item_total // n
            
            cur = con.execute(
                "INSERT INTO purchases(date,model,qty,total,note,user_id) VALUES(?,?,?,?,?,?)",
                (date, item_model, n, item_total, note, uid),
            )
            pid = cur.lastrowid
            pids.append(pid)
            
            for i, s in enumerate(item_serials):
                cost = item_total - base_cost * (n - 1) if i == 0 else base_cost
                con.execute(
                    "INSERT INTO units(serial,model,purchase_id,cost,status,user_id) VALUES(?,?,?,?,?,?)",
                    (s, item_model, pid, cost, status, uid),
                )
        con.commit()
    except sqlite3.IntegrityError:
        con.rollback()
        return bad("貨號已存在")
    except Exception:
        con.rollback()
        raise

    resp = {"ok": True, "ids": pids}
    if len(pids) == 1:
        resp["id"] = pids[0]
    return jsonify(resp)



@app.route("/api/purchase/<int:pid>", methods=["PATCH"])
@auth_required
def edit_purchase(pid):
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    try:
        p = con.execute("SELECT * FROM purchases WHERE id=? AND user_id=?", (pid, uid)).fetchone()
        if not p:
            con.rollback()
            return bad("找不到此筆進貨", 404)
        date = (d.get("date") or p["date"]).strip()
        model = (d.get("model") or p["model"]).strip()
        total = as_int(d.get("total", p["total"]), p["total"])
        note = d.get("note", p["note"])
        if not date or not model or total < 0:
            con.rollback()
            return bad("日期、型號、金額皆為必填")
        if not valid_date(date):
            con.rollback()
            return bad("日期格式須為 YYYY-MM-DD")

        # Optional serials
        serials_input = d.get("serials")
        if serials_input is not None and not isinstance(serials_input, dict):
            con.rollback()
            return bad("貨號資料格式不正確")
        serials_dict = serials_input or {}

        db_units = con.execute("SELECT id, serial FROM units WHERE purchase_id=? AND user_id=?", (pid, uid)).fetchall()
        db_units_dict = {u["id"]: u["serial"] for u in db_units}

        # qty is editable ONLY on ledger-only rows (Excel imports, no linked units);
        # linked purchases derive their count from the machine list.
        qty = p["qty"]
        if "qty" in d:
            new_qty = as_int(d.get("qty"), -1)
            if db_units and new_qty != p["qty"]:
                con.rollback()
                return bad("已連結機器的進貨台數不可修改（刪除後重新登記）")
            if new_qty < 1:
                con.rollback()
                return bad("台數須為 1 以上")
            qty = new_qty

        parsed_serials = {}
        for k, v in serials_dict.items():
            try:
                unit_id = int(k)
            except ValueError:
                con.rollback()
                return bad("機器不屬於此筆進貨")
            if unit_id not in db_units_dict:
                con.rollback()
                return bad("機器不屬於此筆進貨")
            if not isinstance(v, str):
                con.rollback()
                return bad("貨號資料格式不正確")
            new_serial = v.strip()
            if not new_serial:
                con.rollback()
                return bad("貨號不可空白")
            if new_serial != db_units_dict[unit_id]:
                parsed_serials[unit_id] = new_serial

        seen = set()
        for unit_id, s in parsed_serials.items():
            if s in seen:
                con.rollback()
                return bad(f"貨號重複：{s}")
            seen.add(s)

        renamed_ids = list(parsed_serials.keys())
        placeholders = ",".join("?" for _ in renamed_ids)
        for unit_id, s in parsed_serials.items():
            q = f"SELECT 1 FROM units WHERE serial=? AND user_id=? AND id NOT IN ({placeholders})"
            if con.execute(q, [s, uid] + renamed_ids).fetchone():
                con.rollback()
                return bad(f"貨號已存在：{s}")

        con.execute("UPDATE purchases SET date=?, model=?, qty=?, total=?, note=? WHERE id=?",
                    (date, model, qty, total, note, pid))
        if model != p["model"]:
            # keep the machine rows in step (sales keep their denormalized model as history)
            con.execute("UPDATE units SET model=? WHERE purchase_id=? AND user_id=?",
                        (model, pid, uid))
        if total != p["total"]:
            unit_ids = [r["id"] for r in con.execute(
                "SELECT id FROM units WHERE purchase_id=? AND user_id=? ORDER BY id", (pid, uid))]
            n = len(unit_ids)
            if n:
                base = total // n
                for i, u in enumerate(unit_ids):
                    cost = total - base * (n - 1) if i == 0 else base
                    con.execute("UPDATE units SET cost=? WHERE id=?", (cost, u))

        if renamed_ids:
            # First phase: set each renamed unit's serial to a temp value
            for unit_id in renamed_ids:
                temp_serial = "\x00" + str(unit_id)
                con.execute("UPDATE units SET serial=? WHERE id=?", (temp_serial, unit_id))
            # Second phase: set final serials and update sales
            for unit_id, s in parsed_serials.items():
                con.execute("UPDATE units SET serial=? WHERE id=?", (s, unit_id))
                con.execute("UPDATE sales SET serial=? WHERE unit_id=? AND user_id=?", (s, unit_id, uid))

        con.commit()
    except sqlite3.IntegrityError:
        con.rollback()
        return bad("貨號已存在")
    return jsonify(ok=True)


@app.route("/api/purchase/<int:pid>/split", methods=["POST"])
@auth_required
def split_purchase(pid):
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    
    items = d.get("items")
    if not isinstance(items, list) or not (2 <= len(items) <= 12):
        return bad("型號、台數、金額格式不正確")

    validated_items = []
    for item in items:
        if not isinstance(item, dict):
            return bad("型號、台數、金額格式不正確")
        model = (item.get("model") or "").strip()
        qty = as_int(item.get("qty"), 0)
        total = as_int(item.get("total"), -1)
        if not model or qty < 1 or total < 0:
            return bad("型號、台數、金額格式不正確")
        validated_items.append({
            "model": model,
            "qty": qty,
            "total": total
        })

    con = db()
    con.execute("BEGIN IMMEDIATE")
    try:
        p = con.execute("SELECT * FROM purchases WHERE id=? AND user_id=?", (pid, uid)).fetchone()
        if not p:
            con.rollback()
            return bad("找不到此筆進貨", 404)
        
        linked = con.execute("SELECT COUNT(*) FROM units WHERE purchase_id=?", (pid,)).fetchone()[0]
        if linked > 0:
            con.rollback()
            return bad("已連結機器的進貨無法拆單")

        first_item = validated_items[0]
        con.execute(
            "UPDATE purchases SET model=?, qty=?, total=? WHERE id=?",
            (first_item["model"], first_item["qty"], first_item["total"], pid)
        )
        
        new_ids = [pid]
        for item in validated_items[1:]:
            cur = con.execute(
                "INSERT INTO purchases(date,model,qty,total,note,user_id) VALUES(?,?,?,?,?,?)",
                (p["date"], item["model"], item["qty"], item["total"], p["note"], uid)
            )
            new_ids.append(cur.lastrowid)
            
        con.commit()
    except Exception:
        con.rollback()
        raise

    return jsonify(ok=True, ids=new_ids)


@app.route("/api/purchase/<int:pid>", methods=["DELETE"])
@auth_required
def del_purchase(pid):
    uid = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    if not con.execute("SELECT 1 FROM purchases WHERE id=? AND user_id=?", (pid, uid)).fetchone():
        return bad("找不到此筆進貨", 404)
    sold = con.execute(
        "SELECT COUNT(*) FROM units WHERE purchase_id=? AND user_id=? AND status<>'in_stock'",
        (pid, uid)).fetchone()[0]
    if sold:
        return bad("此筆進貨已有機器售出／試用／特許持機／除役，無法刪除")
    con.execute("DELETE FROM units WHERE purchase_id=? AND user_id=?", (pid, uid))
    con.execute("DELETE FROM purchases WHERE id=? AND user_id=?", (pid, uid))
    con.commit()
    return jsonify(ok=True)


# ---------- sales ----------

@app.route("/api/sale", methods=["POST"])
@auth_required
def add_sale():
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    date = (d.get("date") or "").strip()
    customer = (d.get("customer") or "").strip()
    unit_ids = d.get("unit_ids") or []
    if not isinstance(unit_ids, list):
        return bad("貨號資料格式不正確")
    unit_ids = list(dict.fromkeys(unit_ids))
    total_price = as_int(d.get("total_price"), -1)
    card_fee = as_int(d.get("card_fee"), 0)
    warranty = (d.get("warranty_no") or "").strip()
    fixes = d.get("serial_fix") or {}
    if not isinstance(fixes, dict):
        return bad("貨號資料格式不正確")
    note = d.get("note", "")
    sale_type = d.get("sale_type") or "normal"
    if not date or not customer or not unit_ids or total_price < 0:
        return bad("日期、客戶、貨號、金額皆為必填")
    if card_fee < 0 or card_fee > total_price:
        return bad("刷卡手續費格式不正確")
    if not valid_date(date):
        return bad("日期格式須為 YYYY-MM-DD")
    if sale_type not in ("normal", "franchise"):
        return bad("類別不正確")
    extra_fee = as_int(d.get("extra_fee"), 0)
    extra_label = (d.get("extra_label") or "").strip()
    if extra_fee < 0:
        return bad("其他費用格式不正確")
    agent = ""
    deposit = commission = tax = health_fee = 0
    deposit_date = settle_date = ""
    if sale_type == "franchise":
        agent = (d.get("agent") or "").strip()
        if not agent:
            return bad("特許人必填")
        deposit = as_int(d.get("deposit"), -1)
        if deposit < 0:
            return bad("保證金格式不正確")
        if deposit > total_price:
            return bad("保證金不可大於售價")
        deposit_date = (d.get("deposit_date") or "").strip()
        if not deposit_date or not valid_date(deposit_date):
            return bad("保證金收款日格式須為 YYYY-MM-DD")
        commission = total_price - deposit
        # 保證金% + 佣金% = 100%; withholdings come out of the commission, so it may not drop below 12.11%
        if total_price > 0 and commission * 10000 < total_price * 1211:
            return bad("佣金比例不可低於 12.11%")
        # tax/health default to the statutory rates but the form may hand-override them
        tax = as_int(d.get("tax"), -1)
        if tax < 0:
            tax = half_up(commission * WITHHOLD_RATE)
        health_fee = as_int(d.get("health_fee"), -1)
        if health_fee < 0:
            health_fee = half_up(commission * HEALTH_RATE)
        if tax + health_fee > commission:
            return bad("預扣稅款與補充保費合計不可大於佣金")
        # expected payout date (inert until settled=1): next month's 15th unless the form supplies one
        settle_date = (d.get("settle_date") or "").strip()
        if settle_date:
            if not valid_date(settle_date):
                return bad("結清日期格式須為 YYYY-MM-DD")
        else:
            settle_date = next_month_15(date)
    con = db()
    con.execute("BEGIN IMMEDIATE")
    units = [con.execute("SELECT * FROM units WHERE id=? AND user_id=?", (i, uid)).fetchone()
             for i in unit_ids]
    if any(u is None for u in units):
        return bad("找不到指定的機器")
    # franchise sales may sell units a 特許 already holds (consigned); normal sales may not
    ok_status = ("in_stock", "consigned") if sale_type == "franchise" else ("in_stock",)
    not_avail = [u["serial"] for u in units if u["status"] not in ok_status]
    if not_avail:
        consigned = [u["serial"] for u in units if u["status"] == "consigned"]
        if consigned and sale_type != "franchise":
            return bad("特許持機中：" + "、".join(consigned) + "（請用居間特許類別售出）")
        return bad("非在庫（已售／試用機／除役）：" + "、".join(not_avail))
    n = len(units)
    base = total_price // n
    try:
        inserted_ids = []
        for i, u in enumerate(units):
            price = total_price - base * (n - 1) if i == 0 else base
            serial = (fixes.get(str(u["id"])) or "").strip() or u["serial"]
            if serial != u["serial"]:
                if con.execute("SELECT 1 FROM units WHERE serial=? AND user_id=? AND id<>?",
                               (serial, uid, u["id"])).fetchone():
                    return bad("貨號已存在：" + serial)
                con.execute("UPDATE units SET serial=? WHERE id=?", (serial, u["id"]))
            cur = con.execute(
                "INSERT INTO sales(date,customer,unit_id,model,serial,price,card_fee,cost,warranty_no,note,user_id,"
                "sale_type,agent,deposit,deposit_date,commission,tax,health_fee,settled,settle_date,extra_fee,extra_label)"
                " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (date, customer, u["id"], u["model"], serial, price,
                 card_fee if i == 0 else 0, u["cost"], warranty, note, uid,
                 sale_type, agent,
                 deposit if i == 0 else 0, deposit_date if i == 0 else "",
                 commission if i == 0 else 0, tax if i == 0 else 0, health_fee if i == 0 else 0,
                 0, settle_date if i == 0 else "",
                 extra_fee if i == 0 else 0, extra_label if i == 0 else ""),
            )
            inserted_ids.append(cur.lastrowid)
            con.execute("UPDATE units SET status='sold' WHERE id=?", (u["id"],))
            # a sold consigned unit consumes its ACTIVE 特許領機 record — its deposit info
            # now lives on the sale row (keeps dep_in counted exactly once). Returned
            # consignments (returned=1) are historical and must be preserved.
            con.execute("DELETE FROM consignments WHERE unit_id=? AND user_id=? AND returned=0", (u["id"], uid))
        if inserted_ids:
            first_id = inserted_ids[0]
            placeholders = ",".join("?" for _ in inserted_ids)
            con.execute(f"UPDATE sales SET group_id=? WHERE id IN ({placeholders})", [first_id] + inserted_ids)
        con.commit()
    except sqlite3.IntegrityError:
        return bad("貨號已存在")
    return jsonify(ok=True)


@app.route("/api/sale/<int:sid>", methods=["PATCH"])
@auth_required
def edit_sale(sid):
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    s = con.execute("SELECT * FROM sales WHERE id=? AND user_id=?", (sid, uid)).fetchone()
    if not s:
        return bad("找不到此筆銷售", 404)
    if s["group_id"] is not None:
        cnt = con.execute("SELECT COUNT(*) FROM sales WHERE group_id=? AND user_id=?", (s["group_id"], uid)).fetchone()[0]
        if cnt > 1:
            return bad("此筆屬多台交易，請重新整理頁面後以整組方式編輯")
    date = (d.get("date") or s["date"]).strip()
    customer = (d.get("customer") or s["customer"]).strip()
    model = (d.get("model") or s["model"]).strip()
    serial = ((d["serial"] if "serial" in d else s["serial"]) or "").strip()
    price = as_int(d.get("price", s["price"]), s["price"])
    card_fee = as_int(d.get("card_fee", s["card_fee"]), s["card_fee"])
    cost = as_int(d.get("cost", s["cost"]), s["cost"])
    warranty = ((d["warranty_no"] if "warranty_no" in d else s["warranty_no"]) or "").strip()
    note = d.get("note", s["note"])
    sale_type = d.get("sale_type") or s["sale_type"]
    if sale_type not in ("normal", "franchise"):
        return bad("類別不正確")
    agent = ((d["agent"] if "agent" in d else s["agent"]) or "").strip()
    deposit = as_int(d.get("deposit", s["deposit"]), s["deposit"])
    commission = as_int(d.get("commission", s["commission"]), s["commission"])
    tax = as_int(d.get("tax", s["tax"]), s["tax"])
    health_fee = as_int(d.get("health_fee", s["health_fee"]), s["health_fee"])
    deposit_date = ((d["deposit_date"] if "deposit_date" in d else s["deposit_date"]) or "").strip()
    settle_date = ((d["settle_date"] if "settle_date" in d else s["settle_date"]) or "").strip()
    settled = 1 if d.get("settled", s["settled"]) else 0
    extra_fee = as_int(d.get("extra_fee", s["extra_fee"]), s["extra_fee"])
    extra_label = ((d["extra_label"] if "extra_label" in d else s["extra_label"]) or "").strip()
    if extra_fee < 0:
        return bad("其他費用格式不正確")
    if not date or not customer or price < 0 or card_fee < 0 or cost < 0:
        return bad("日期、客戶、金額皆為必填")
    if card_fee < 0 or card_fee > price:
        return bad("刷卡手續費格式不正確")
    if not valid_date(date):
        return bad("日期格式須為 YYYY-MM-DD")
    if deposit < 0 or commission < 0 or tax < 0 or health_fee < 0:
        return bad("金額格式不正確")
    if (deposit_date and not valid_date(deposit_date)) or (settle_date and not valid_date(settle_date)):
        return bad("日期格式須為 YYYY-MM-DD")
    if settled and not settle_date:
        settle_date = datetime.date.today().isoformat()
    # a settled franchise row is a closed cycle: freeze every material field (category,
    # P&L amounts, franchise money, dates) until the owner explicitly un-settles first
    # (pass settled=0). This blocks the sale_type->normal force-zero bypass too.
    if s["settled"] == 1 and s["sale_type"] == "franchise" and settled:
        if (sale_type != s["sale_type"] or
            price != s["price"] or cost != s["cost"] or card_fee != s["card_fee"] or
            extra_fee != s["extra_fee"] or
            deposit != s["deposit"] or commission != s["commission"] or
            tax != s["tax"] or health_fee != s["health_fee"] or
            deposit_date != s["deposit_date"] or settle_date != s["settle_date"]):
            return bad("已結清，請先將此筆改為未結清再修改金額、成本或類別")
    if sale_type == "franchise":
        if not agent:
            return bad("特許人必填")
        # deal value = 保證金＋佣金 (= the WHOLE deal's price). On multi-unit sales the row's
        # price is only a per-unit share while the money sits on the first row, so per-row
        # price must not be the ceiling here (it 400'd every edit of such rows). Zero-money
        # sibling rows (deal=0) are exempt from the money rules (incl. deposit_date, which
        # also lives on the first row only); a fat-fingered deposit still trips the floor.
        deal = deposit + commission
        if deal > 0 and not deposit_date:
            return bad("保證金收款日為必填")
        if deal > 0 and commission * 10000 < deal * 1211:
            return bad("佣金比例不可低於 12.11%")
        if tax + health_fee > commission:
            return bad("預扣稅款與補充保費合計不可大於佣金")
    else:
        # normal rows must carry no franchise money — MONTHLY_SQL sums commission unconditionally
        agent, deposit_date, settle_date = "", "", ""
        deposit = commission = tax = health_fee = settled = 0
    try:
        if s["unit_id"]:
            if not serial:
                return bad("此筆已連結機器，貨號不可空白")
            if serial != s["serial"]:
                if con.execute("SELECT 1 FROM units WHERE serial=? AND user_id=? AND id<>?",
                               (serial, uid, s["unit_id"])).fetchone():
                    return bad("貨號已存在")
                con.execute("UPDATE units SET serial=? WHERE id=?", (serial, s["unit_id"]))
        con.execute(
            "UPDATE sales SET date=?, customer=?, model=?, serial=?, price=?, card_fee=?,"
            " cost=?, warranty_no=?, note=?, sale_type=?, agent=?, deposit=?, deposit_date=?, commission=?,"
            " tax=?, health_fee=?, settled=?, settle_date=?, extra_fee=?, extra_label=? WHERE id=?",
            (date, customer, model, serial, price, card_fee, cost, warranty, note,
             sale_type, agent, deposit, deposit_date, commission, tax, health_fee, settled, settle_date,
             extra_fee, extra_label, sid),
        )
        con.commit()
    except sqlite3.IntegrityError:
        return bad("貨號已存在")
    return jsonify(ok=True)


@app.route("/api/sale/<int:sid>", methods=["DELETE"])
@auth_required
def del_sale(sid):
    uid = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    row = con.execute("SELECT * FROM sales WHERE id=? AND user_id=?", (sid, uid)).fetchone()
    if not row:
        return bad("找不到此筆銷售", 404)
    if row["group_id"] is not None:
        cnt = con.execute("SELECT COUNT(*) FROM sales WHERE group_id=? AND user_id=?", (row["group_id"], uid)).fetchone()[0]
        if cnt > 1:
            return bad("此筆屬多台交易，請重新整理頁面後以整組方式編輯")
    # a settled franchise sale is a closed cycle (deposit refunded + commission paid);
    # deleting it would drop the payout record AND wrongly re-create a live consignment.
    if row["sale_type"] == "franchise" and row["settled"] == 1:
        return bad("已結清的居間特許無法刪除，請先於編輯中改為未結清")
    if row["unit_id"]:
        if row["sale_type"] == "franchise" and row["deposit"] > 0:
            con.execute(
                "INSERT INTO consignments(agent,unit_id,deposit,deposit_date,note,user_id)"
                " VALUES(?,?,?,?,?,?)",
                (row["agent"], row["unit_id"], row["deposit"], row["deposit_date"], row["note"], uid)
            )
            con.execute("UPDATE units SET status='consigned' WHERE id=? AND user_id=?", (row["unit_id"], uid))
        else:
            con.execute("UPDATE units SET status='in_stock' WHERE id=? AND status='sold'",
                        (row["unit_id"],))
    con.execute("DELETE FROM sales WHERE id=?", (sid,))
    con.commit()
    return jsonify(ok=True)


@app.route("/api/sale-group/<int:gid>", methods=["PATCH"])
@auth_required
def edit_sale_group(gid):
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    rows = con.execute("SELECT * FROM sales WHERE group_id=? AND user_id=?", (gid, uid)).fetchall()
    if not rows:
        return bad("找不到此筆銷售群組", 404)
    
    anchor = next((r for r in rows if r["id"] == gid), None)
    if not anchor:
        return bad("資料異常：找不到主列", 400)
    
    date = (d.get("date") or anchor["date"])
    if isinstance(date, str):
        date = date.strip()
    customer = (d.get("customer") or anchor["customer"])
    if isinstance(customer, str):
        customer = customer.strip()
    sale_type = d.get("sale_type") or anchor["sale_type"]
    agent = d.get("agent") if "agent" in d else anchor["agent"]
    if agent is None:
        agent = ""
    agent = agent.strip()
    note = d.get("note") if "note" in d else anchor["note"]
    if note is None:
        note = ""
    
    if not date or not customer:
        return bad("日期、客戶皆為必填")
    if not valid_date(date):
        return bad("日期格式須為 YYYY-MM-DD")
    if sale_type not in ("normal", "franchise"):
        return bad("類別不正確")
        
    card_fee = as_int(d.get("card_fee", anchor["card_fee"]), anchor["card_fee"])
    extra_fee = as_int(d.get("extra_fee", anchor["extra_fee"]), anchor["extra_fee"])
    extra_label = d.get("extra_label") if "extra_label" in d else anchor["extra_label"]
    if extra_label is None:
        extra_label = ""
    extra_label = extra_label.strip()
    warranty = d.get("warranty_no") if "warranty_no" in d else anchor["warranty_no"]
    if warranty is None:
        warranty = ""
    warranty = warranty.strip()
    deposit = as_int(d.get("deposit", anchor["deposit"]), anchor["deposit"])
    commission = as_int(d.get("commission", anchor["commission"]), anchor["commission"])
    tax = as_int(d.get("tax", anchor["tax"]), anchor["tax"])
    health_fee = as_int(d.get("health_fee", anchor["health_fee"]), anchor["health_fee"])
    deposit_date = d.get("deposit_date") if "deposit_date" in d else anchor["deposit_date"]
    if deposit_date is None:
        deposit_date = ""
    deposit_date = deposit_date.strip()
    settle_date = d.get("settle_date") if "settle_date" in d else anchor["settle_date"]
    if settle_date is None:
        settle_date = ""
    settle_date = settle_date.strip()
    settled = 1 if d.get("settled", anchor["settled"]) else 0
    
    if card_fee < 0:
        return bad("刷卡手續費格式不正確")
    if extra_fee < 0:
        return bad("其他費用格式不正確")
    if deposit < 0 or commission < 0 or tax < 0 or health_fee < 0:
        return bad("金額格式不正確")
    if (deposit_date and not valid_date(deposit_date)) or (settle_date and not valid_date(settle_date)):
        return bad("日期格式須為 YYYY-MM-DD")
    if settled and not settle_date:
        settle_date = datetime.date.today().isoformat()
        
    # Prices handling
    total_price_payload = d.get("total_price")
    prices_map = d.get("prices")
    if total_price_payload is not None and prices_map is not None:
        return bad("金額參數重複", 400)
        
    n = len(rows)
    new_prices = {}
    if total_price_payload is not None:
        total_price = as_int(total_price_payload, -1)
        if total_price < 0:
            return bad("總價格式不正確")
        base = total_price // n
        for r in rows:
            if r["id"] == gid:
                new_prices[r["id"]] = total_price - base * (n - 1)
            else:
                new_prices[r["id"]] = base
    elif prices_map is not None:
        if not isinstance(prices_map, dict):
            return bad("金額分配格式不正確")
        row_ids_in_group = {r["id"] for r in rows}
        try:
            parsed_prices_map = {int(k): as_int(v, -1) for k, v in prices_map.items()}
        except ValueError:
            return bad("金額分配格式不正確")
        if set(parsed_prices_map.keys()) != row_ids_in_group:
            return bad("金額分配必須包含群組中所有機器且不能有多餘項目")
        if any(v < 0 for v in parsed_prices_map.values()):
            return bad("每台機器的售價不可小於 0")
        new_prices = parsed_prices_map
        total_price = sum(new_prices.values())
    else:
        for r in rows:
            new_prices[r["id"]] = r["price"]
        total_price = sum(new_prices.values())
        
    if card_fee > total_price:
        return bad("刷卡手續費格式不正確")
        
    # Costs handling
    costs_map = d.get("costs") or {}
    if not isinstance(costs_map, dict):
        return bad("成本分配格式不正確")
    for rid_str, c_val in costs_map.items():
        try:
            rid = int(rid_str)
            c_int = as_int(c_val, -1)
            if c_int < 0:
                return bad("成本不可小於 0")
        except ValueError:
            return bad("成本分配格式不正確")
            
    # SETTLED-FREEZE group-wide
    any_settled_franchise = any(r["settled"] == 1 and r["sale_type"] == "franchise" for r in rows)
    if any_settled_franchise and settled:
        price_changed = any(new_prices[r["id"]] != r["price"] for r in rows)
        cost_changed = False
        for rid_str, new_c_val in costs_map.items():
            try:
                rid = int(rid_str)
                r_stored = next(r for r in rows if r["id"] == rid)
                if as_int(new_c_val, r_stored["cost"]) != r_stored["cost"]:
                    cost_changed = True
            except ValueError:
                pass
        if (sale_type != anchor["sale_type"] or
            price_changed or cost_changed or
            card_fee != anchor["card_fee"] or
            extra_fee != anchor["extra_fee"] or
            deposit != anchor["deposit"] or
            commission != anchor["commission"] or
            tax != anchor["tax"] or
            health_fee != anchor["health_fee"] or
            deposit_date != anchor["deposit_date"] or
            settle_date != anchor["settle_date"]):
            return bad("已結清，請先將此筆改為未結清再修改金額、成本或類別")
            
    if sale_type == "franchise":
        if not agent:
            return bad("特許人必填")
        deal = deposit + commission
        if deal > 0 and not deposit_date:
            return bad("保證金收款日為必填")
        if deal > 0 and total_price > 0 and commission * 10000 < deal * 1211:
            return bad("佣金比例不可低於 12.11%")
        if tax + health_fee > commission:
            return bad("預扣稅款與補充保費合計不可大於佣金")
    else:
        agent, deposit_date, settle_date = "", "", ""
        deposit = commission = tax = health_fee = settled = 0
        
    # Serials renaming
    serials_dict = d.get("serials") or {}
    if not isinstance(serials_dict, dict):
        return bad("貨號資料格式不正確")
    
    parsed_serials = {}
    row_ids_in_group = {r["id"] for r in rows}
    for k, v in serials_dict.items():
        try:
            sale_id = int(k)
        except ValueError:
            return bad("群組內找不到指定的銷售列")
        if sale_id not in row_ids_in_group:
            return bad("群組內找不到指定的銷售列")
        if not isinstance(v, str):
            return bad("貨號資料格式不正確")
        new_serial = v.strip()
        if not new_serial:
            return bad("貨號不可空白")
        r_stored = next(r for r in rows if r["id"] == sale_id)
        if new_serial != r_stored["serial"]:
            parsed_serials[sale_id] = new_serial
            
    seen = set()
    for r in rows:
        if r["id"] not in parsed_serials:
            seen.add(r["serial"])
    for sale_id, s_val in parsed_serials.items():
        if s_val in seen:
            return bad(f"貨號重複：{s_val}")
        seen.add(s_val)
        
    renamed_unit_ids = [r["unit_id"] for r in rows if r["id"] in parsed_serials and r["unit_id"]]
    if renamed_unit_ids:
        placeholders = ",".join("?" for _ in renamed_unit_ids)
        for sale_id, s_val in parsed_serials.items():
            r_stored = next(r for r in rows if r["id"] == sale_id)
            if r_stored["unit_id"]:
                q = f"SELECT 1 FROM units WHERE serial=? AND user_id=? AND id NOT IN ({placeholders})"
                if con.execute(q, [s_val, uid] + renamed_unit_ids).fetchone():
                    return bad(f"貨號已存在：{s_val}")
                    
    try:
        # Phase 1 of serial update
        for sale_id, s_val in parsed_serials.items():
            r_stored = next(r for r in rows if r["id"] == sale_id)
            u_id = r_stored["unit_id"]
            if u_id:
                temp_serial = "\x00" + str(u_id)
                con.execute("UPDATE units SET serial=? WHERE id=?", (temp_serial, u_id))
        
        # Phase 2 of serial update
        for sale_id, s_val in parsed_serials.items():
            r_stored = next(r for r in rows if r["id"] == sale_id)
            u_id = r_stored["unit_id"]
            if u_id:
                con.execute("UPDATE units SET serial=? WHERE id=?", (s_val, u_id))
                
        # Main updates
        for r in rows:
            rid = r["id"]
            is_anchor = (rid == gid)
            r_price = new_prices[rid]
            r_cost = as_int(costs_map.get(str(rid), r["cost"]), r["cost"])
            r_serial = parsed_serials.get(rid, r["serial"])
            
            con.execute(
                "UPDATE sales SET date=?, customer=?, sale_type=?, agent=?, note=?, model=?,"
                " price=?, cost=?, serial=?,"
                " card_fee=?, extra_fee=?, extra_label=?, warranty_no=?, deposit=?, commission=?, tax=?, health_fee=?, deposit_date=?,"
                " settled=?, settle_date=?"
                " WHERE id=?",
                (
                    date, customer, sale_type, agent, note, r["model"],
                    r_price, r_cost, r_serial,
                    card_fee if is_anchor else 0,
                    extra_fee if is_anchor else 0,
                    extra_label if is_anchor else "",
                    warranty if is_anchor else "",
                    deposit if is_anchor else 0,
                    commission if is_anchor else 0,
                    tax if is_anchor else 0,
                    health_fee if is_anchor else 0,
                    deposit_date if is_anchor else "",
                    settled,
                    # keep the (預計) settle_date even while unsettled — v16 semantics: it's
                    # pre-planned and inert until settled=1; only the anchor carries it
                    settle_date if is_anchor else "",
                    rid
                )
            )
        con.commit()
    except sqlite3.IntegrityError:
        return bad("貨號已存在")
    return jsonify(ok=True)


@app.route("/api/sale-group/<int:gid>", methods=["DELETE"])
@auth_required
def delete_sale_group(gid):
    uid = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    rows = con.execute("SELECT * FROM sales WHERE group_id=? AND user_id=?", (gid, uid)).fetchall()
    if not rows:
        return bad("找不到此筆銷售群組", 404)
    anchor = next((r for r in rows if r["id"] == gid), None)
    if not anchor:
        return bad("資料異常：找不到主列", 400)
    
    if any(r["sale_type"] == "franchise" and r["settled"] == 1 for r in rows):
        return bad("已結清的居間特許無法刪除，請先於編輯中改為未結清")
        
    for r in rows:
        u_id = r["unit_id"]
        if not u_id:
            continue
        if r["id"] == gid and anchor["sale_type"] == "franchise" and anchor["deposit"] > 0:
            con.execute(
                "INSERT INTO consignments(agent,unit_id,deposit,deposit_date,note,user_id)"
                " VALUES(?,?,?,?,?,?)",
                (anchor["agent"], u_id, anchor["deposit"], anchor["deposit_date"], anchor["note"], uid)
            )
            con.execute("UPDATE units SET status='consigned' WHERE id=? AND user_id=?", (u_id, uid))
        else:
            con.execute("UPDATE units SET status='in_stock' WHERE id=? AND status='sold'", (u_id,))
            
    con.execute("DELETE FROM sales WHERE group_id=? AND user_id=?", (gid, uid))
    con.commit()
    return jsonify(ok=True)


@app.route("/api/sale-group/<int:gid>/settle", methods=["POST"])
@auth_required
def settle_sale_group(gid):
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    rows = con.execute("SELECT * FROM sales WHERE group_id=? AND user_id=?", (gid, uid)).fetchall()
    if not rows:
        return bad("找不到此筆銷售群組", 404)
    anchor = next((r for r in rows if r["id"] == gid), None)
    if not anchor:
        return bad("資料異常：找不到主列", 400)
        
    if all(r["settled"] == 1 for r in rows):
        return bad("已結清", 400)
        
    settle_date = (d.get("settle_date") or "").strip()
    if settle_date:
        if not valid_date(settle_date):
            return bad("結清日期格式須為 YYYY-MM-DD")
    else:
        settle_date = anchor["settle_date"] or datetime.date.today().isoformat()
        
    con.execute(
        "UPDATE sales SET settled=1, settle_date=? WHERE group_id=? AND user_id=?",
        (settle_date, gid, uid)
    )
    con.commit()
    return jsonify(ok=True)


@app.route("/api/settle-agent", methods=["POST"])
@auth_required
def settle_agent():
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    agent = (d.get("agent") or "").strip()
    if not agent:
        return bad("特許人必填")
    settle_date = (d.get("settle_date") or "").strip()
    if settle_date:
        if not valid_date(settle_date):
            return bad("結清日期格式須為 YYYY-MM-DD")
    else:
        settle_date = datetime.date.today().isoformat()
    con = db()
    cur = con.execute(
        "UPDATE sales SET settled=1, settle_date=? "
        "WHERE user_id=? AND sale_type='franchise' AND agent=? AND settled=0",
        (settle_date, uid, agent),
    )
    con.commit()
    return jsonify(ok=True, count=cur.rowcount)


# ---------- trials ----------

@app.route("/api/trial", methods=["POST"])
@auth_required
def add_trial():
    d = request.get_json(silent=True) or {}
    customer = (d.get("customer") or "").strip()
    if not customer:
        return bad("請填寫人名")
    start = (d.get("start_date") or "").strip()
    end = (d.get("end_date") or "").strip()
    for dv in (start, end):
        if dv and not valid_date(dv):
            return bad("日期格式須為 YYYY-MM-DD")
    if start and end and end < start:
        return bad("結束日不可早於開始日")
    rent_type = (d.get("rent_type") or "").strip()
    if rent_type and rent_type not in RENT_TYPES:
        return bad("租類不正確")
    con = db()
    con.execute(
        "INSERT INTO trials(customer,model,start_date,end_date,note,user_id,rent_type) VALUES(?,?,?,?,?,?,?)",
        (customer, (d.get("model") or ""), start, end,
         (d.get("note") or ""), g.data_uid, rent_type),
    )
    con.commit()
    return jsonify(ok=True)


@app.route("/api/trial/<int:tid>/return", methods=["POST"])
@auth_required
def return_trial(tid):
    d = request.get_json(silent=True) or {}
    ret_date = d.get("return_date", "")
    if ret_date:
        ret_date = ret_date.strip()
        if not valid_date(ret_date):
            return bad("日期格式須為 YYYY-MM-DD")
    else:
        ret_date = datetime.date.today().isoformat()
    con = db()
    con.execute("UPDATE trials SET returned=1, return_date=? WHERE id=? AND user_id=?", (ret_date, tid, g.data_uid))
    con.commit()
    return jsonify(ok=True)


@app.route("/api/trial/<int:tid>", methods=["PATCH"])
@auth_required
def edit_trial(tid):
    d = request.get_json(silent=True) or {}
    con = db()
    t = con.execute("SELECT * FROM trials WHERE id=? AND user_id=?", (tid, g.data_uid)).fetchone()
    if not t:
        return bad("找不到此筆試用", 404)
    customer = ((d["customer"] if "customer" in d else t["customer"]) or "").strip()
    model = ((d["model"] if "model" in d else t["model"]) or "").strip()
    start = ((d["start_date"] if "start_date" in d else t["start_date"]) or "").strip()
    end = ((d["end_date"] if "end_date" in d else t["end_date"]) or "").strip()
    if (start and not valid_date(start)) or (end and not valid_date(end)):
        return bad("日期格式須為 YYYY-MM-DD")
    if start and end and end < start:
        return bad("結束日不可早於開始日")
    note = d.get("note", t["note"])
    returned = 1 if d.get("returned", t["returned"]) else 0
    rent_type = ((d["rent_type"] if "rent_type" in d else t["rent_type"]) or "").strip()
    if rent_type and rent_type not in RENT_TYPES:
        return bad("租類不正確")
    ret_date = ((d["return_date"] if "return_date" in d else t["return_date"]) or "").strip()
    if ret_date and not valid_date(ret_date):
        return bad("日期格式須為 YYYY-MM-DD")
    if returned == 0:
        ret_date = ""
    con.execute(
        "UPDATE trials SET customer=?, model=?, start_date=?, end_date=?, note=?, returned=?, rent_type=?, return_date=? WHERE id=?",
        (customer, model, start, end, note, returned, rent_type, ret_date, tid),
    )
    con.commit()
    return jsonify(ok=True)


@app.route("/api/trial/<int:tid>", methods=["DELETE"])
@auth_required
def del_trial(tid):
    con = db()
    con.execute("DELETE FROM trials WHERE id=? AND user_id=?", (tid, g.data_uid))
    con.commit()
    return jsonify(ok=True)


# ---------- consignments (特許領機) ----------

@app.route("/api/consign", methods=["POST"])
@auth_required
def add_consign():
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    agent = (d.get("agent") or "").strip()
    if not agent:
        return bad("特許人必填")
    deposit = as_int(d.get("deposit"), -1)
    if deposit < 0:
        return bad("保證金格式不正確")
    deposit_date = (d.get("deposit_date") or "").strip()
    if not deposit_date or not valid_date(deposit_date):
        return bad("保證金收款日格式須為 YYYY-MM-DD")
    con = db()
    con.execute("BEGIN IMMEDIATE")   # atomic availability check + status flip across threads
    u = con.execute("SELECT * FROM units WHERE id=? AND user_id=?",
                    (d.get("unit_id"), uid)).fetchone()
    if not u:
        return bad("找不到指定的機器")
    if u["status"] != "in_stock":
        return bad("非在庫（已售／試用機／除役／特許持機）：" + u["serial"])
    cur = con.execute(
        "INSERT INTO consignments(agent,unit_id,deposit,deposit_date,note,user_id)"
        " VALUES(?,?,?,?,?,?)",
        (agent, u["id"], deposit, deposit_date, d.get("note", ""), uid))
    con.execute("UPDATE units SET status='consigned' WHERE id=?", (u["id"],))
    con.commit()
    return jsonify(ok=True, id=cur.lastrowid)


@app.route("/api/consign/<int:cid>", methods=["PATCH"])
@auth_required
def edit_consign(cid):
    d = request.get_json(silent=True) or {}
    uid = g.data_uid
    con = db()
    cg = con.execute("SELECT * FROM consignments WHERE id=? AND user_id=?", (cid, uid)).fetchone()
    if not cg:
        return bad("找不到此筆特許領機", 404)
    if cg["returned"]:
        return bad("已退回的特許領機無法修改（金流紀錄已結案）")
    agent = ((d["agent"] if "agent" in d else cg["agent"]) or "").strip()
    if not agent:
        return bad("特許人必填")
    deposit = as_int(d.get("deposit", cg["deposit"]), cg["deposit"])
    if deposit < 0:
        return bad("保證金格式不正確")
    deposit_date = ((d["deposit_date"] if "deposit_date" in d else cg["deposit_date"]) or "").strip()
    if not deposit_date or not valid_date(deposit_date):
        return bad("保證金收款日格式須為 YYYY-MM-DD")
    note = d.get("note", cg["note"])
    con.execute("UPDATE consignments SET agent=?, deposit=?, deposit_date=?, note=? WHERE id=?",
                (agent, deposit, deposit_date, note, cid))
    con.commit()
    return jsonify(ok=True)


@app.route("/api/consign/<int:cid>/return", methods=["POST"])
@auth_required
def return_consign(cid):
    uid = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    cg = con.execute("SELECT * FROM consignments WHERE id=? AND user_id=?", (cid, uid)).fetchone()
    if not cg:
        return bad("找不到此筆特許領機", 404)
    if cg["returned"]:
        return bad("此筆已退回")
    d = request.get_json(silent=True) or {}
    refund_date = (d.get("refund_date") or "").strip()
    if not refund_date:
        refund_date = datetime.date.today().isoformat()
    if not valid_date(refund_date):
        return bad("退回日期格式須為 YYYY-MM-DD")
    ref_amt_val = d.get("refund_amount")
    if ref_amt_val is None:
        refund_amount = cg["deposit"]
    else:
        refund_amount = as_int(ref_amt_val, -1)
        if refund_amount < 0:
            return bad("退款金額格式不正確")
    con.execute("UPDATE consignments SET returned=1, refund_date=?, refund_amount=? WHERE id=?",
                (refund_date, refund_amount, cid))
    if cg["unit_id"]:
        con.execute("UPDATE units SET status='in_stock' WHERE id=? AND status='consigned'",
                    (cg["unit_id"],))
    con.commit()
    return jsonify(ok=True)


@app.route("/api/consign/<int:cid>", methods=["DELETE"])
@auth_required
def del_consign(cid):
    uid = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    cg = con.execute("SELECT * FROM consignments WHERE id=? AND user_id=?", (cid, uid)).fetchone()
    if not cg:
        return bad("找不到此筆特許領機", 404)
    if cg["unit_id"] and not cg["returned"]:
        con.execute("UPDATE units SET status='in_stock' WHERE id=? AND status='consigned'",
                    (cg["unit_id"],))
    con.execute("DELETE FROM consignments WHERE id=?", (cid,))
    con.commit()
    return jsonify(ok=True)


# ---------- units ----------

@app.route("/api/unit/<int:uid_>", methods=["PATCH"])
@auth_required
def edit_unit(uid_):
    d = request.get_json(silent=True) or {}
    owner = g.data_uid
    con = db()
    con.execute("BEGIN IMMEDIATE")
    u = con.execute("SELECT * FROM units WHERE id=? AND user_id=?", (uid_, owner)).fetchone()
    if not u:
        return bad("找不到機器", 404)
    serial = ((d.get("serial") or u["serial"]) or "").strip()
    note = d.get("note", u["note"])
    cost = as_int(d.get("cost", u["cost"]), u["cost"])
    if cost < 0:
        return bad("成本不可小於 0")
    status = d.get("status", u["status"])
    if status not in UNIT_STATUSES:
        return bad("狀態不正確")
    if u["status"] == "sold" and status != "sold":
        return bad("已售出的機器請先刪除該筆銷售")
    if u["status"] != "sold" and status == "sold":
        return bad("請用「銷售」登記售出")
    if u["status"] == "consigned" and status != "consigned":
        return bad("特許持機中，請先於銷售頁售出或取消該筆特許領機")
    if u["status"] != "consigned" and status == "consigned":
        return bad("請用「特許領機」登記")
    dup = con.execute("SELECT 1 FROM units WHERE serial=? AND user_id=? AND id<>?",
                      (serial, owner, uid_)).fetchone()
    if dup:
        return bad("貨號已存在")
    try:
        con.execute("UPDATE units SET serial=?, note=?, cost=?, status=? WHERE id=?",
                    (serial, note, cost, status, uid_))
        con.execute("UPDATE sales SET serial=? WHERE unit_id=?", (serial, uid_))
        con.commit()
    except sqlite3.IntegrityError:
        return bad("貨號已存在")
    return jsonify(ok=True)


# ---------- export ----------

def xl(v):
    return ("'" + v) if isinstance(v, str) and v[:1] in ("=", "+", "-", "@") else v


def build_workbook(con, uid):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", start_color="3B4A9F")
    money = "#,##0"

    def style_head(ws, ncols, widths):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center")
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws = wb.active
    ws.title = "月報"
    ws.append(["月份", "銷售總額", "銷貨成本", "佣金", "其他費用", "毛利", "銷售數量", "毛利率"])
    rows = list(con.execute(MONTHLY_SQL, (uid,)))
    for r in rows:
        ws.append([r["ym"], r["revenue"], r["cost"], r["commission"], r["extra"], r["profit"], r["qty"],
                   (r["profit"] / r["revenue"]) if r["revenue"] else 0])
    if rows:
        tr = sum(r["revenue"] for r in rows)
        tc = sum(r["cost"] for r in rows)
        tcm = sum(r["commission"] for r in rows)
        tx = sum(r["extra"] for r in rows)
        tq = sum(r["qty"] for r in rows)
        tp = tr - tc - tcm - tx
        ws.append(["合計", tr, tc, tcm, tx, tp, tq, (tp / tr) if tr else 0])
        ws.cell(row=len(rows) + 2, column=1).font = Font(bold=True)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
        for cell in row:
            cell.number_format = money
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = "0.0%"
    style_head(ws, 8, [10, 12, 12, 11, 11, 12, 10, 9])

    ws = wb.create_sheet("特許金流")
    ws.append(["月份", "保證金收", "退保證金", "實付佣金", "預扣稅款", "補充保費", "淨流"])
    frows = list(con.execute(FRANCHISE_FLOW_SQL, (uid, uid, uid, uid)))
    for r in frows:
        net = r["dep_in"] - r["dep_out"] - r["comm_net"] - r["tax"] - r["health"]
        ws.append([r["ym"], r["dep_in"], r["dep_out"], r["comm_net"], r["tax"], r["health"], net])
    if frows:
        tdi = sum(r["dep_in"] for r in frows)
        tdo = sum(r["dep_out"] for r in frows)
        tcn = sum(r["comm_net"] for r in frows)
        ttx = sum(r["tax"] for r in frows)
        thl = sum(r["health"] for r in frows)
        ws.append(["合計", tdi, tdo, tcn, ttx, thl, tdi - tdo - tcn - ttx - thl])
        ws.cell(row=len(frows) + 2, column=1).font = Font(bold=True)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=7):
        for cell in row:
            cell.number_format = money
    style_head(ws, 7, [10, 12, 12, 12, 12, 12, 12])

    ws = wb.create_sheet("特許人扣繳彙總")
    ws.append(["年度", "特許人", "筆數", "佣金合計", "預扣稅款合計", "補充保費合計", "實付佣金合計"])
    q_tax = """
    SELECT substr(settle_date,1,4) AS yr, agent,
           COUNT(*) AS n, SUM(commission) AS comm, SUM(tax) AS tax, SUM(health_fee) AS health,
           SUM(commission - tax - health_fee) AS net
    FROM sales WHERE user_id=? AND sale_type='franchise' AND agent<>''
      AND settled=1 AND settle_date<>''
    GROUP BY yr, agent ORDER BY yr, agent
    """
    for r in con.execute(q_tax, (uid,)):
        ws.append([r["yr"], xl(r["agent"]), r["n"], r["comm"], r["tax"], r["health"], r["net"]])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=7):
        for cell in row:
            cell.number_format = money
    style_head(ws, 7, [8, 12, 8, 12, 12, 12, 12])

    ws = wb.create_sheet("銷售明細")
    ws.append(["日期", "客戶", "類別", "特許人", "型號", "貨號", "保證書編號", "銷售單價", "刷卡費",
               "其他費用", "費用名稱", "實收", "進貨成本", "毛利", "保證金", "保證金收款日", "佣金",
               "預扣稅款", "補充保費", "實付佣金", "結清", "結清日期", "備註"])
    for s in con.execute("SELECT * FROM sales WHERE user_id=? ORDER BY date, id", (uid,)):
        net = s["price"] - s["card_fee"]
        is_fr = s["sale_type"] == "franchise"
        category = "居間特許" if is_fr else "一般"
        gp = net - s["cost"] - s["commission"] - s["extra_fee"]
        money_row = is_fr and (s["deposit"] > 0 or s["commission"] > 0)
        if money_row:
            franchise_cells = [s["deposit"], s["deposit_date"], s["commission"], s["tax"], s["health_fee"],
                                s["commission"] - s["tax"] - s["health_fee"],
                                "已結清" if s["settled"] else "未結清", s["settle_date"]]
        else:
            franchise_cells = ["", "", "", "", "", "", "", ""]
        ws.append([s["date"], xl(s["customer"]), category, xl(s["agent"]), xl(s["model"]), xl(s["serial"]),
                   xl(s["warranty_no"]), s["price"], s["card_fee"],
                   s["extra_fee"] or "", xl(s["extra_label"]), net, s["cost"], gp,
                   *franchise_cells, xl(s["note"])])
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=10):
        for cell in row:
            cell.number_format = money
    for row in ws.iter_rows(min_row=2, min_col=12, max_col=15):
        for cell in row:
            cell.number_format = money
    for row in ws.iter_rows(min_row=2, min_col=17, max_col=20):
        for cell in row:
            cell.number_format = money
    style_head(ws, 23, [11, 10, 10, 10, 11, 12, 13, 11, 9, 10, 11, 11, 11, 11, 11, 14, 10, 11, 11, 11, 8, 12, 24])

    ws = wb.create_sheet("進貨明細")
    ws.append(["日期", "型號", "數量", "金額", "備註"])
    for p in con.execute("SELECT * FROM purchases WHERE user_id=? ORDER BY date, id", (uid,)):
        ws.append([p["date"], xl(p["model"]), p["qty"], p["total"], xl(p["note"])])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = money
    style_head(ws, 5, [11, 24, 8, 12, 36])

    ws = wb.create_sheet("庫存")
    ws.append(["貨號", "型號", "狀態", "成本", "備註"])
    label = {"in_stock": "在庫", "sold": "已售", "trial": "試用機", "retired": "除役", "consigned": "特許機"}
    for u in con.execute("SELECT * FROM units WHERE user_id=? ORDER BY status, model, serial", (uid,)):
        ws.append([xl(u["serial"]), xl(u["model"]), label.get(u["status"], u["status"]), u["cost"], xl(u["note"])])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = money
    style_head(ws, 5, [14, 12, 9, 11, 32])

    ws = wb.create_sheet("試用出租")
    ws.append(["人名", "型號", "租類", "歸還日", "開始", "結束", "狀態", "備註"])
    rent_label_map = {"week7": "七天租", "month": "月租", "franchise": "特許租用", "hq": "總部月租", "reserve": "預約", "": ""}
    for t in con.execute("SELECT * FROM trials WHERE user_id=? ORDER BY returned, start_date", (uid,)):
        ws.append([xl(t["customer"]), xl(t["model"]), rent_label_map.get(t["rent_type"], ""), xl(t["return_date"]), t["start_date"], t["end_date"],
                   "已歸還" if t["returned"] else "進行中", xl(t["note"])])
    style_head(ws, 8, [12, 11, 11, 11, 11, 11, 9, 28])
    return wb


def build_tax_workbook(con, uid, year):
    from copy import copy
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # Fill the accountant's real 範本 in place so the styling matches it exactly
    # (標楷體, borders, row heights, print layout — nothing rebuilt by hand).
    wb = load_workbook(os.path.join(BASE, "tax_template.xlsx"))
    ws = wb.active
    ws["A1"] = f"{year - 1911}年度 執行業務所得印領清冊 "

    # settled franchise payouts keyed by 結清日 (= 給付日). Columns: H..S = 一月..十二月
    # of the selected year, T=合計. G (前期佣金) is left BLANK for the owners to hand-fill.
    q = """
    SELECT agent, substr(settle_date,1,7) AS ym,
           SUM(commission) AS comm, SUM(tax) AS tax, SUM(health_fee) AS health,
           SUM(commission - tax - health_fee) AS net
    FROM sales WHERE user_id=? AND sale_type='franchise' AND settled=1 AND agent<>'' AND settle_date<>''
      AND substr(settle_date,1,7) BETWEEN ? AND ?
    GROUP BY agent, ym
    ORDER BY agent, ym
    """
    rows = con.execute(q, (uid, f"{year}-01", f"{year}-12")).fetchall()
    month_col = {}
    for m in range(1, 13):
        month_col[f"{year}-{m:02d}"] = 7 + m
    agents = {}
    for r in rows:
        agents.setdefault(r["agent"], {})[r["ym"]] = r
    names = sorted(agents)

    TPL_SLOTS = 5   # the 範本 ships five pre-styled 4-row blocks (rows 3..22, totals at 23)
    BLOCK0, TOTALS0 = 3, 23
    extra = max(0, len(names) - TPL_SLOTS)
    if extra:
        # grow the sheet: 4 rows per extra agent above the totals block, cloning the
        # first block's cell styles; the shifted totals formulas are rewritten below.
        ws.insert_rows(TOTALS0, 4 * extra)
        for k in range(extra):
            base = TOTALS0 + 4 * k
            for off in range(4):
                for col in range(1, 21):
                    ws.cell(row=base + off, column=col)._style = copy(
                        ws.cell(row=BLOCK0 + off, column=col)._style)
                ws.row_dimensions[base + off].height = ws.row_dimensions[BLOCK0 + off].height
            for col in "ABCDE":
                ws.merge_cells(f"{col}{base}:{col}{base + 3}")
            ws.cell(row=base, column=1, value=TPL_SLOTS + k + 1)
            for off, lbl in enumerate(("佣金", "扣繳稅額10%", "補充保費2.11%", "實領")):
                ws.cell(row=base + off, column=6, value=lbl)
            for off in range(4):
                ws.cell(row=base + off, column=20, value=f"=SUM(G{base + off}:S{base + off})")
        totals_row = TOTALS0 + 4 * extra
        for off in range(4):
            for c_idx in range(7, 21):
                col = get_column_letter(c_idx)
                ws.cell(row=totals_row + off, column=c_idx,
                        value="=" + "+".join(f"{col}{BLOCK0 + 4 * k + off}" for k in range(len(names))))

    for idx, name in enumerate(names):
        r0 = BLOCK0 + 4 * idx
        ws.cell(row=r0, column=3, value=xl(name))
        per = agents[name]
        for ym, col in month_col.items():
            if ym in per:
                rec = per[ym]
                for off, v in enumerate((rec["comm"], rec["tax"], rec["health"], rec["net"])):
                    ws.cell(row=r0 + off, column=col, value=v)
    return wb


@app.route("/api/export.xlsx")
@auth_required
def export_xlsx():
    wb = build_workbook(db(), g.data_uid)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = "DENBA_進銷存_" + datetime.date.today().strftime("%Y%m%d") + ".xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=name,
    )


@app.route("/api/tax-export.xlsx")
@auth_required
def tax_export_xlsx():
    year_param = request.args.get("year", "")
    current_year = datetime.date.today().year
    year = as_int(year_param, current_year)
    if not (2000 <= year <= 2100):
        return bad("年度不正確")
    
    wb = build_tax_workbook(db(), g.data_uid, year)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    roc_year = year - 1911
    name = f"執行業務所得清冊_{roc_year}年度.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=name,
    )


# ---------- backups / restore ----------
# Two layers:
#   1. per-user JSON snapshots — any user can back up / restore THEIR OWN rows
#   2. whole-DB file snapshots (cron + pre-restore) — admin only, affects everyone

SYS_BK_RE = re.compile(r"^denba-(\d{8}|pre-restore-\d{8}-\d{6})\.db$")
USER_BK_RE = re.compile(r"^user(\d+)-(pre-restore-|pre-delete-)?(\d{8}-\d{6})\.json$")


def snapshot_db(dest_path):
    src = sqlite3.connect(DB_PATH)
    src.execute("PRAGMA busy_timeout=5000")
    dst = sqlite3.connect(dest_path)
    dst.execute("PRAGMA busy_timeout=5000")
    src.backup(dst)
    dst.close()
    src.close()


def dump_user(con, uid):
    out = {}
    for t in DATA_TABLES:
        out[t] = [dict(r) for r in con.execute(f"SELECT * FROM {t} WHERE user_id=?", (uid,))]
    return out


def write_user_snapshot(con, uid, tag=""):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    prefix = f"user{uid}-{tag}-" if tag else f"user{uid}-"
    name = f"{prefix}{ts}.json"
    with open(os.path.join(BACKUP_DIR, name), "w", encoding="utf-8") as f:
        json.dump(dump_user(con, uid), f, ensure_ascii=False)
    keep = 15
    mine = sorted(f for f in os.listdir(BACKUP_DIR)
                  if USER_BK_RE.match(f) and USER_BK_RE.match(f).group(1) == str(uid))
    for old in mine[:-keep]:
        os.unlink(os.path.join(BACKUP_DIR, old))
    return name


def restore_user(con, uid, payload):
    for t in DATA_TABLES:
        con.execute(f"DELETE FROM {t} WHERE user_id=?", (uid,))
    pmap, umap = {}, {}
    for p in payload.get("purchases", []):
        cur = con.execute(
            "INSERT INTO purchases(date,model,qty,total,note,user_id) VALUES(?,?,?,?,?,?)",
            (p["date"], p["model"], p["qty"], p["total"], p.get("note", ""), uid))
        pmap[p["id"]] = cur.lastrowid
    for u in payload.get("units", []):
        cur = con.execute(
            "INSERT INTO units(serial,model,purchase_id,cost,status,note,user_id)"
            " VALUES(?,?,?,?,?,?,?)",
            (u["serial"], u["model"], pmap.get(u.get("purchase_id")), u["cost"],
             u["status"], u.get("note", ""), uid))
        umap[u["id"]] = cur.lastrowid
    for cg in payload.get("consignments", []):
        con.execute(
            "INSERT INTO consignments(agent,unit_id,deposit,deposit_date,note,user_id,"
            "returned,refund_date,refund_amount)"
            " VALUES(?,?,?,?,?,?,?,?,?)",
            (cg.get("agent", ""), umap.get(cg.get("unit_id")), cg.get("deposit", 0),
             cg.get("deposit_date", ""), cg.get("note", ""), uid,
             cg.get("returned", 0), cg.get("refund_date", ""), cg.get("refund_amount", 0)))
    smap = {}
    sales_to_update = []
    for s in payload.get("sales", []):
        cur = con.execute(
            "INSERT INTO sales(date,customer,unit_id,model,serial,price,card_fee,cost,warranty_no,note,user_id,"
            "sale_type,agent,deposit,deposit_date,commission,tax,health_fee,settled,settle_date,extra_fee,extra_label,group_id)"
            " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (s["date"], s["customer"], umap.get(s.get("unit_id")), s["model"],
             s.get("serial", ""), s["price"], s.get("card_fee", 0), s["cost"],
             s.get("warranty_no", ""), s.get("note", ""), uid,
             s.get("sale_type", "normal"), s.get("agent", ""), s.get("deposit", 0),
             s.get("deposit_date", ""), s.get("commission", 0), s.get("tax", 0),
             s.get("health_fee", 0), s.get("settled", 0), s.get("settle_date", ""),
             s.get("extra_fee", 0), s.get("extra_label", ""), None))
        new_id = cur.lastrowid
        if "id" in s:
            smap[s["id"]] = new_id
            if s.get("group_id") is not None:
                sales_to_update.append((new_id, s["group_id"]))
    for new_id, old_group_id in sales_to_update:
        new_group_id = smap.get(old_group_id)
        con.execute("UPDATE sales SET group_id=? WHERE id=?", (new_group_id, new_id))
    for t in payload.get("trials", []):
        con.execute(
            "INSERT INTO trials(customer,model,start_date,end_date,note,returned,user_id,rent_type,return_date)"
            " VALUES(?,?,?,?,?,?,?,?,?)",
            (t.get("customer", ""), t.get("model", ""), t.get("start_date", ""),
             t.get("end_date", ""), t.get("note", ""), t.get("returned", 0), uid,
             t.get("rent_type", ""), t.get("return_date", "")))


@app.route("/api/backups")
@auth_required
def list_backups():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    uid = str(g.data_uid)
    user_backups, system_backups = [], []
    for fn in os.listdir(BACKUP_DIR):
        full = os.path.join(BACKUP_DIR, fn)
        m = USER_BK_RE.match(fn)
        if m and m.group(1) == uid:
            st = os.stat(full)
            user_backups.append({"name": fn, "size": st.st_size, "mtime": int(st.st_mtime)})
        elif SYS_BK_RE.match(fn) and g.user["is_admin"]:
            st = os.stat(full)
            system_backups.append({"name": fn, "size": st.st_size, "mtime": int(st.st_mtime)})
    user_backups.sort(key=lambda x: x["mtime"], reverse=True)
    system_backups.sort(key=lambda x: (x["mtime"], x["name"]), reverse=True)
    return jsonify(user_backups=user_backups, system_backups=system_backups)


@app.route("/api/backup-now", methods=["POST"])
@auth_required
def backup_now():
    con = db()
    name = write_user_snapshot(con, g.data_uid)
    con.commit()
    return jsonify(ok=True, name=name)


@app.route("/api/restore", methods=["POST"])
@auth_required
def restore_backup():
    name = (request.get_json(silent=True) or {}).get("name", "")
    con = db()
    um = USER_BK_RE.match(name)
    if um:
        if um.group(1) != str(g.data_uid):
            return bad("只能還原自己的備份", 403)
        path = os.path.join(BACKUP_DIR, name)
        if not os.path.exists(path):
            return bad("找不到備份檔", 404)
        with open(path, encoding="utf-8") as f:
            payload = json.load(f)
        pre = write_user_snapshot(con, g.data_uid, "pre-restore")
        restore_user(con, g.data_uid, payload)
        con.commit()
        return jsonify(ok=True, pre_restore=pre)
    if SYS_BK_RE.match(name):
        if not g.user["is_admin"]:
            return bad("需要管理員權限", 403)
        path = os.path.join(BACKUP_DIR, name)
        if not os.path.exists(path):
            return bad("找不到備份檔", 404)
        pre = "denba-pre-restore-" + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + ".db"
        snapshot_db(os.path.join(BACKUP_DIR, pre))
        src = sqlite3.connect(path)
        src.execute("PRAGMA busy_timeout=5000")
        dst = sqlite3.connect(DB_PATH)
        dst.execute("PRAGMA busy_timeout=5000")
        src.backup(dst)
        dst.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        dst.close()
        src.close()
        pres = sorted(f for f in os.listdir(BACKUP_DIR) if f.startswith("denba-pre-restore-"))
        for f in pres[:-10]:
            os.unlink(os.path.join(BACKUP_DIR, f))
        return jsonify(ok=True, pre_restore=pre)
    return bad("備份檔名不正確")


init_db()

if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "export":
        cli_uid = int(sys.argv[3]) if len(sys.argv) >= 4 else 1
        cli_con = sqlite3.connect(DB_PATH)
        cli_con.execute("PRAGMA busy_timeout=5000")
        cli_con.row_factory = sqlite3.Row
        build_workbook(cli_con, cli_uid).save(sys.argv[2])
        cli_con.close()
    else:
        if not APP_PASSWORD:
            raise SystemExit("APP_PASSWORD 未設定（請在 denba.env 設定）")
        if not os.environ.get("SECRET_KEY"):
            raise SystemExit("SECRET_KEY 未設定（請在 denba.env 設定）")
        from waitress import serve
        serve(app, host="0.0.0.0", port=PORT, threads=4)
